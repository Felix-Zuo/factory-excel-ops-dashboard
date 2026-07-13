"""Small spreadsheet readers used by the sample workflow.

CSV is supported without dependencies. XLSX is supported through openpyxl when
installed. The reader returns dictionaries keyed by the header row.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


def read_table(path: Path) -> list[dict[str, object]]:
    """Read a CSV or XLSX file into a list of dictionaries."""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")


def read_headers(path: Path, max_rows: int = 25) -> list[str]:
    """Return likely headers from the first non-empty row."""

    rows = read_table_preview(path, max_rows=max_rows)
    return list(rows[0].keys()) if rows else []


def read_table_preview(path: Path, max_rows: int = 25) -> list[dict[str, object]]:
    """Read at most max_rows rows."""

    return read_table(path)[:max_rows]


def _read_csv(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to read XLSX files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(_iter_rows(sheet.iter_rows(values_only=True)))
    if not rows:
        return []

    header_index = _find_header_row(rows)
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    result: list[dict[str, object]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        record = {
            header: value
            for header, value in zip(headers, row)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            record["_row_number"] = row_number
            result.append(record)
    workbook.close()
    return result


def _iter_rows(rows: Iterable[tuple[object, ...]]) -> Iterable[tuple[object, ...]]:
    for row in rows:
        if any(value not in (None, "") for value in row):
            yield row


def _find_header_row(rows: list[tuple[object, ...]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        score = sum(1 for value in row if isinstance(value, str) and value.strip())
        if score > best_score:
            best_index = index
            best_score = score
    return best_index
